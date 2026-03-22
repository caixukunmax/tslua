#!/usr/bin/env python
"""
清理 __tables__.xlsx 中所有包含 test/ 的行
"""

import openpyxl
import sys
from pathlib import Path

# 文件路径
file_path = Path(__file__).parent / 'datas' / '__tables__.xlsx'

print(f"正在处理文件：{file_path}")

# 加载工作簿
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"处理前总行数：{ws.max_row}")

# 找到 input 列的索引（假设第一行是表头）
header = [cell.value for cell in ws[1]]
try:
    input_col_idx = header.index('input') + 1  # openpyxl 列索引从 1 开始
    print(f"input 列索引：{input_col_idx} (列 {openpyxl.utils.get_column_letter(input_col_idx)})")
except ValueError:
    print("错误：找不到 'input' 列")
    print(f"可用列：{header}")
    sys.exit(1)

# 收集需要删除的行（从下往上，避免索引错乱）
rows_to_delete = []
for row in range(ws.max_row, 1, -1):  # 从最后一行到第 2 行（跳过表头）
    cell_value = ws.cell(row=row, column=input_col_idx).value
    if cell_value and 'test/' in str(cell_value):
        rows_to_delete.append(row)
        print(f"  标记删除行 {row}: {cell_value}")

# 删除标记的行
for row_num in rows_to_delete:
    ws.delete_rows(row_num)

print(f"删除了 {len(rows_to_delete)} 行")
print(f"处理后总行数：{ws.max_row}")

# 保存文件
wb.save(file_path)
print(f"✓ 文件已保存：{file_path}")
