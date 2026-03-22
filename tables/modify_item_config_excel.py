#!/usr/bin/env python
"""
修改道具配置表 Excel 文件（基于现有文件模板）
"""

import openpyxl
from pathlib import Path

file_path = Path(__file__).parent / 'datas' / 'item' / '道具配置表.xlsx'
print(f"正在修改：{file_path}")

# 加载工作簿
wb = openpyxl.load_workbook(file_path)

# 重命名 Sheet
ws = wb.active
ws.title = "道具配置表"

# 取消所有合并单元格
merged_ranges = list(ws.merged_cells.ranges)
for merged_range in merged_ranges:
    ws.unmerge_cells(merged_range.coord)

# 清空所有单元格
for row in ws.iter_rows():
    for cell in row:
        cell.value = None

# 定义新表头
headers = [
    ("id", "int"),
    ("name", "string"),
    ("major_type", "EMajorType"),
    ("minor_type", "EMinorType"),
    ("quality", "EItemQuality"),
    ("max_pile_num", "int"),
    ("icon", "string"),
    ("icon_background", "string"),
    ("icon_mask", "string"),
    ("desc", "string"),
    ("show_order", "int"),
    ("effect_type", "string"),
    ("effect_value", "int"),
    ("price", "int"),
    ("can_sell", "bool"),
    ("obtain_methods", "string"),
    ("release_date", "string"),
]

# Luban 格式：
# R1: ## | field1 | field2 | ...
# R2: ## | comment1 | comment2 | ...  (可选的中文注释)
# R3: ## | type_comment1 | type_comment2 | ...  (可选的类型说明)
# R4: data_row1
# R5: data_row2
# ...

# 第 1 行：## + 字段名
ws.cell(row=1, column=1, value="##")
for col, (field, _) in enumerate(headers, 1):
    ws.cell(row=1, column=col+1, value=field)

# 第 2 行：## + 空（或中文注释）
ws.cell(row=2, column=1, value="##")
# 留空，不添加额外注释

# 第 3 行：## + 空
ws.cell(row=3, column=1, value="##")
# 留空，不添加类型注释

# 示例数据（使用正确的枚举值）
sample_data = [
    [1001, "生命药水", "CONSUMABLES", "FOOD", "GREEN", 99, "icons/item_1001.png", "bg_green", "", "恢复少量生命值", 10, "heal", 50, 100, True, "副本掉落，商店购买", "2024-01-01"],
    [1002, "魔法药水", "CONSUMABLES", "FOOD", "BLUE", 99, "icons/item_1002.png", "bg_blue", "", "恢复少量魔法值", 11, "mana", 30, 150, True, "副本掉落，任务奖励", "2024-01-01"],
    [1003, "高级生命药水", "CONSUMABLES", "FOOD", "PURPLE", 50, "icons/item_1003.png", "bg_purple", "", "恢复中量生命值", 12, "heal", 200, 300, True, "精英副本，活动兑换", "2024-01-01"],
    [1004, "传说生命药水", "CONSUMABLES", "FOOD", "GOLDEN", 20, "icons/item_1004.png", "bg_golden", "mask_rare", "恢复大量生命值，稀有道具", 13, "heal", 1000, 1000, True, "世界 BOSS，限时活动", "2024-01-15"],
    [1005, "钻石", "CURRENCY", "DIAMOND", "GOLDEN", 9999, "icons/diamond.png", "bg_orange", "", "premium 货币", 1, "gem", 1, 100, False, "充值，成就奖励", "2024-01-01"],
    [1006, "金币", "CURRENCY", "GOLD", "WHITE", 99999, "icons/gold.png", "bg_yellow", "", "基础货币", 2, "gold", 1, 1, False, "所有玩法", "2024-01-01"],
    [1007, "经验书·小", "CONSUMABLES", "EXP", "WHITE", 999, "icons/exp_small.png", "bg_green", "", "少量经验值", 14, "exp", 100, 50, True, "日常任务", "2024-01-01"],
    [1008, "经验书·大", "CONSUMABLES", "EXP", "BLUE", 999, "icons/exp_large.png", "bg_blue", "", "大量经验值", 15, "exp", 1000, 200, True, "周常任务，活动", "2024-01-01"],
    [1009, "强化石", "CONSUMABLES", "CONSTRUCTION_MATERIAL", "GREEN", 999, "icons/enhance_stone.png", "bg_gray", "", "装备强化材料", 16, "material", 1, 100, True, "分解装备，副本", "2024-01-01"],
    [1010, "设计图纸·剑", "CONSUMABLES", "DESIGN_DRAWING", "PURPLE", 1, "icons/design_sword.png", "bg_purple", "mask_common", "武器制作图纸", 17, "recipe", 1, 500, True, "工匠等级奖励", "2024-01-20"],
]

# 写入数据（从第 4 行开始）
for row_idx, data in enumerate(sample_data, 4):
    # 第 1 列（A 列）填"是"表示启用
    ws.cell(row=row_idx, column=1, value="是")
    # 从第 2 列（B 列）开始填数据
    for col_idx, value in enumerate(data, 2):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 保存
wb.save(file_path)
print(f"✓ 修改成功")
print(f"  - Sheet 名：{ws.title}")
print(f"  - {len(headers)} 个字段")
print(f"  - {len(sample_data)} 条示例数据")
