#!/usr/bin/env python
"""
清理 __tables__.xlsx - 删除所有 test/相关的行
"""

import openpyxl
from pathlib import Path

file_path = Path(__file__).parent / 'datas' / '__tables__.xlsx'
print(f"正在处理：{file_path}")

# 加载工作簿
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 找到 input 列（E 列）
input_col = 5  # E column

# 收集需要删除的行（包含 test/ 的行）
rows_to_delete = []
for row in range(ws.max_row, 1, -1):  # 从下往上
    cell_value = ws.cell(row=row, column=input_col).value
    if cell_value and 'test/' in str(cell_value):
        rows_to_delete.append(row)
        print(f"  标记删除行 {row}: {cell_value}")

# 删除行
for row_num in rows_to_delete:
    ws.delete_rows(row_num)

# 保存
wb.save(file_path)
print(f"\n✓ 已删除 {len(rows_to_delete)} 行")
print(f"✓ 剩余行数：{ws.max_row}")
