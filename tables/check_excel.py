#!/usr/bin/env python
"""
检查 Excel 文件中引用的类型
"""

import openpyxl
from pathlib import Path

# 文件路径
file_path = Path(__file__).parent / 'datas' / '#AutoImport1.xlsx'

print(f"正在检查文件：{file_path}")

# 加载工作簿
wb = openpyxl.load_workbook(file_path)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    print(f"\n=== 工作表：{ws_name} ===")
    
    # 读取表头
    header = [cell.value for cell in ws[1]]
    print(f"列：{header}")
    
    # 查找包含 test. 的单元格
    found = False
    for row in range(2, min(ws.max_row + 1, 20)):  # 检查前 20 行
        for col in range(1, len(header) + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and 'test.' in str(cell_value):
                print(f"  行{row}, 列{header[col-1]}: {cell_value}")
                found = True
    
    if not found:
        print("  ✓ 未发现 test. 引用")
