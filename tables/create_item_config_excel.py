#!/usr/bin/env python
"""
创建道具配置表 Excel 文件
"""

import openpyxl
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 文件路径
file_path = Path(__file__).parent / 'datas' / 'item' / '道具配置表.xlsx'

print(f"正在创建文件：{file_path}")

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "道具表"

# 定义样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头（根据 Luban 定义）
headers = [
    ("id", "int", "道具 ID"),
    ("name", "string", "道具名称"),
    ("major_type", "EMajorType", "主类型"),
    ("minor_type", "EMinorType", "子类型"),
    ("quality", "EItemQuality", "品质"),
    ("max_pile_num", "int", "最大堆叠数"),
    ("icon", "string", "图标路径"),
    ("icon_background", "string", "图标背景"),
    ("icon_mask", "string", "图标遮罩"),
    ("desc", "string", "描述"),
    ("show_order", "int", "显示顺序"),
    ("effect_type", "string", "效果类型"),
    ("effect_value", "int", "效果数值"),
    ("price", "int", "价格"),
    ("can_sell", "bool", "可否出售"),
    ("obtain_methods", "string", "获取途径"),
    ("release_date", "string", "上线日期"),
]

# 写入表头
for col, (field_name, field_type, comment) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=field_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    
    # 添加注释行
    ws.cell(row=2, column=col, value=f"##type:{field_type}")
    
    # 第三行写示例数据
    if col == 1:  # id
        ws.cell(row=3, column=col, value=1001)
    elif col == 2:  # name
        ws.cell(row=3, column=col, value="生命药水")
    elif col == 3:  # major_type
        ws.cell(row=3, column=col, value="CONSUMABLES")
    elif col == 4:  # minor_type
        ws.cell(row=3, column=col, value="FOOD")
    elif col == 5:  # quality
        ws.cell(row=3, column=col, value="GREEN")
    elif col == 6:  # max_pile_num
        ws.cell(row=3, column=col, value=99)
    elif col == 7:  # icon
        ws.cell(row=3, column=col, value="icons/item_1001.png")
    elif col == 8:  # icon_background
        ws.cell(row=3, column=col, value="bg_green")
    elif col == 9:  # icon_mask
        ws.cell(row=3, column=col, value="")
    elif col == 10:  # desc
        ws.cell(row=3, column=col, value="恢复少量生命值")
    elif col == 11:  # show_order
        ws.cell(row=3, column=col, value=10)
    elif col == 12:  # effect_type
        ws.cell(row=3, column=col, value="heal")
    elif col == 13:  # effect_value
        ws.cell(row=3, column=col, value=50)
    elif col == 14:  # price
        ws.cell(row=3, column=col, value=100)
    elif col == 15:  # can_sell
        ws.cell(row=3, column=col, value=True)
    elif col == 16:  # obtain_methods
        ws.cell(row=3, column=col, value="副本掉落，商店购买")
    elif col == 17:  # release_date
        ws.cell(row=3, column=col, value="2024-01-01")

# 添加更多示例数据
sample_data = [
    [1002, "魔法药水", "CONSUMABLES", "FOOD", "BLUE", 99, "icons/item_1002.png", "bg_blue", "", "恢复少量魔法值", 11, "mana", 30, 150, True, "副本掉落，任务奖励", "2024-01-01"],
    [1003, "高级生命药水", "CONSUMABLES", "FOOD", "PURPLE", 50, "icons/item_1003.png", "bg_purple", "", "恢复中量生命值", 12, "heal", 200, 300, True, "精英副本，活动兑换", "2024-01-01"],
    [1004, "传说生命药水", "CONSUMABLES", "FOOD", "GOLDEN", 20, "icons/item_1004.png", "bg_golden", "mask_rare", "恢复大量生命值，稀有道具", 13, "heal", 1000, 1000, True, "世界 BOSS，限时活动", "2024-01-15"],
    [1005, "钻石", "CURRENCY", "DIAMOND", "GOLDEN", 9999, "icons/diamond.png", "bg_orange", "", "premium 货币", 1, "gem", 1, 100, False, "充值，成就奖励", "2024-01-01"],
    [1006, "金币", "CURRENCY", "GOLD", "WHITE", 99999, "icons/gold.png", "bg_yellow", "", "基础货币", 2, "gold", 1, 1, False, "所有玩法", "2024-01-01"],
    [1007, "经验书·小", "CONSUMABLES", "MATERIAL", "WHITE", 999, "icons/exp_small.png", "bg_green", "", "少量经验值", 14, "exp", 100, 50, True, "日常任务", "2024-01-01"],
    [1008, "经验书·大", "CONSUMABLES", "MATERIAL", "BLUE", 999, "icons/exp_large.png", "bg_blue", "", "大量经验值", 15, "exp", 1000, 200, True, "周常任务，活动", "2024-01-01"],
    [1009, "强化石", "MATERIAL", "CONSTRUCTION_MATERIAL", "GREEN", 999, "icons/enhance_stone.png", "bg_gray", "", "装备强化材料", 16, "material", 1, 100, True, "分解装备，副本", "2024-01-01"],
    [1010, "设计图纸·剑", "MATERIAL", "DESIGN_DRAWING", "PURPLE", 1, "icons/design_sword.png", "bg_purple", "mask_common", "武器制作图纸", 17, "recipe", 1, 500, True, "工匠等级奖励", "2024-01-20"],
]

for row_idx, data in enumerate(sample_data, 4):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# 调整列宽
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# 保存文件
wb.save(file_path)
print(f"✓ 文件已创建：{file_path}")
print(f"  - 共 {len(sample_data) + 3} 行数据（包含表头和示例）")
print(f"  - 字段数：{len(headers)}")
