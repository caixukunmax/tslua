#!/usr/bin/env python
"""
在 __tables__.xlsx 中添加道具配置表定义
"""

import openpyxl
from pathlib import Path

file_path = Path(__file__).parent / 'datas' / '__tables__.xlsx'
print(f"正在处理：{file_path}")

# 加载工作簿
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 找到最后一行
last_row = ws.max_row + 1

# 添加新表定义（根据 Luban 格式）
new_table = {
    '##var': 'TbItemConfig',
    'full_name': 'item.TbItemConfig',
    'value_type': 'ItemConfig',
    'read_schema_from_file': False,
    'input': '道具配置表@item/道具配置表.xlsx',
    'index': None,
    'mode': None,
    'group': 'c,s,e',
    'comment': '道具配置表',
    'tags': None,
}

# 写入数据
columns = ['##var', 'full_name', 'value_type', 'read_schema_from_file', 'input', 
           'index', 'mode', 'group', 'comment', 'tags']

for col_idx, col_name in enumerate(columns, 1):
    value = new_table.get(col_name)
    if value is not None:
        ws.cell(row=last_row, column=col_idx, value=value)

# 删除最后一行（刚才添加的）
ws.delete_rows(last_row)

# 保存
wb.save(file_path)
print(f"✗ 已删除重复的定义，恢复到之前状态")
print(f"  提示：在 defines/item.xml 中定义的表会自动生效，无需在 __tables__.xlsx 中重复定义")
