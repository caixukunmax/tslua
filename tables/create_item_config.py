#!/usr/bin/env python
"""
创建道具配置表 Excel 文件 - TbItemConfig
"""

import openpyxl
from pathlib import Path

# 文件路径
file_path = Path(__file__).parent / 'datas' / 'item' / '道具配置表.xlsx'
print(f"正在创建：{file_path}")

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "道具配置表"  # Sheet 名必须与 Luban 定义一致

# 表头定义 (字段名，类型)
headers = [
    ("id", "int"),
    ("name", "string"),
    ("major_type", "EMajorType"),
    ("minor_type", "EMinorType"),
    ("quality", "EItemQuality"),
    ("max_pile_num", "int"),
    ("icon", "string"),
    ("icon_background", "string"),
    ("icon_mask", "string"),
    ("desc", "string"),
    ("show_order", "int"),
    ("effect_type", "string"),
    ("effect_value", "int"),
    ("price", "int"),
    ("can_sell", "bool"),
    ("obtain_methods", "string"),
    ("release_date", "string"),
]

# 第 1 行：Luban 特殊标记 ##
ws.cell(row=1, column=1, value="##")

# 第 2 行：字段名
for col, (field, _) in enumerate(headers, 1):
    ws.cell(row=2, column=col, value=field)

# 第 3 行：类型注释
for col, (_, ftype) in enumerate(headers, 1):
    ws.cell(row=3, column=col, value=f"##type:{ftype}")

# 示例数据（10 个道具）
sample_data = [
    # id, name, major_type, minor_type, quality, max_pile, icon, icon_bg, icon_mask, desc, order, effect_type, effect_val, price, can_sell, obtain, date
    [1001, "生命药水", "CONSUMABLES", "FOOD", "GREEN", 99, "icons/item_1001.png", "bg_green", "", "恢复少量生命值", 10, "heal", 50, 100, True, "副本掉落，商店购买", "2024-01-01"],
    [1002, "魔法药水", "CONSUMABLES", "FOOD", "BLUE", 99, "icons/item_1002.png", "bg_blue", "", "恢复少量魔法值", 11, "mana", 30, 150, True, "副本掉落，任务奖励", "2024-01-01"],
    [1003, "高级生命药水", "CONSUMABLES", "FOOD", "PURPLE", 50, "icons/item_1003.png", "bg_purple", "", "恢复中量生命值", 12, "heal", 200, 300, True, "精英副本，活动兑换", "2024-01-01"],
    [1004, "传说生命药水", "CONSUMABLES", "FOOD", "GOLDEN", 20, "icons/item_1004.png", "bg_golden", "mask_rare", "恢复大量生命值，稀有道具", 13, "heal", 1000, 1000, True, "世界 BOSS，限时活动", "2024-01-15"],
    [1005, "钻石", "CURRENCY", "DIAMOND", "GOLDEN", 9999, "icons/diamond.png", "bg_orange", "", "premium 货币", 1, "gem", 1, 100, False, "充值，成就奖励", "2024-01-01"],
    [1006, "金币", "CURRENCY", "GOLD", "WHITE", 99999, "icons/gold.png", "bg_yellow", "", "基础货币", 2, "gold", 1, 1, False, "所有玩法", "2024-01-01"],
    [1007, "经验书·小", "CONSUMABLES", "MATERIAL", "WHITE", 999, "icons/exp_small.png", "bg_green", "", "少量经验值", 14, "exp", 100, 50, True, "日常任务", "2024-01-01"],
    [1008, "经验书·大", "CONSUMABLES", "MATERIAL", "BLUE", 999, "icons/exp_large.png", "bg_blue", "", "大量经验值", 15, "exp", 1000, 200, True, "周常任务，活动", "2024-01-01"],
    [1009, "强化石", "MATERIAL", "CONSTRUCTION_MATERIAL", "GREEN", 999, "icons/enhance_stone.png", "bg_gray", "", "装备强化材料", 16, "material", 1, 100, True, "分解装备，副本", "2024-01-01"],
    [1010, "设计图纸·剑", "MATERIAL", "DESIGN_DRAWING", "PURPLE", 1, "icons/design_sword.png", "bg_purple", "mask_common", "武器制作图纸", 17, "recipe", 1, 500, True, "工匠等级奖励", "2024-01-20"],
]

# 写入数据（从第 4 行开始）
for row_idx, data in enumerate(sample_data, 4):
    for col_idx, value in enumerate(data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 调整列宽
column_widths = [12, 15, 15, 15, 12, 12, 25, 15, 15, 30, 10, 12, 12, 10, 10, 25, 12]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# 保存
wb.save(file_path)
print(f"✓ 创建成功：{file_path}")
print(f"  - {len(headers)} 个字段")
print(f"  - {len(sample_data)} 条示例数据")
